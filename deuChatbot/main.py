from langchain.text_splitter import RecursiveCharacterTextSplitter, CharacterTextSplitter, KonlpyTextSplitter
from langchain_openai import OpenAIEmbeddings
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_upstage import UpstageEmbeddings
from langchain_chroma import Chroma
from langchain_openai import ChatOpenAI
from langchain_anthropic import ChatAnthropic
from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler

from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.prompts import ChatPromptTemplate

from langchain.schema.runnable import RunnableLambda, RunnablePassthrough

import os
import shutil
from dotenv import load_dotenv
import openpyxl
from openpyxl import load_workbook

import numpy as np
import pandas as pd

from langchain_community.document_loaders import TextLoader
from langchain_community.retrievers import BM25Retriever
from langchain.retrievers import EnsembleRetriever
from langchain_community.document_transformers import LongContextReorder
from langchain_core.output_parsers import StrOutputParser

from langchain.chains.query_constructor.base import AttributeInfo
from langchain.retrievers.self_query.base import SelfQueryRetriever

from typing import List, Tuple
import time

from kiwipiepy import Kiwi
from konlpy.tag import Kkma, Okt

from implements.preprocessing import GPTScorePreprocessing


class ChatBotSystem:
    def __init__(self):
        self.system = "ChatBotSystem"
        self.kiwi = Kiwi()
        self.kkma = Kkma()
        self.okt = Okt()

    def load_env(self):
        load_dotenv('.env')

        os.getenv("LANGCHAIN_TRACING_V2")
        os.getenv("LANGCHAIN_ENDPOINT")
        os.getenv("LANGCHAIN_API_KEY")

        os.getenv("OPENAI_API_KEY")
        os.getenv("ANTHROPIC_API_KEY")
        os.getenv("GOOGLE_API_KEY")
        os.getenv("UPSTAGE_API_KEY")

        os.getenv("LM_URL")
        os.getenv("LM_LOCAL_URL")

    def docs_load(self) -> List[str]:
        """
        문서를 읽는 함수
        """

        try:
            # loader = TextLoader("corpus/모집요강 전처리 버전 1.txt", encoding="utf-8").load()
            # loader = TextLoader("corpus/모집요강 전처리 버전 2.txt", encoding="utf-8").load()
            # loader = TextLoader("corpus/모집요강 전처리 버전 3.txt", encoding="utf-8").load()
            loader = TextLoader("corpus/모집요강 전처리 버전 4.txt", encoding="utf-8").load()
            return loader
        except FileNotFoundError:
            print("파일을 찾을 수 없습니다. 경로를 확인하세요.")
            return []
        except Exception as e:
            print(f"오류가 발생했습니다: {e}")
            return []

    def c_text_split(self, corpus: List[str]) -> List[str]:
        """
        CharacterTextSplitter를 사용하여 문서를 분할하도록 하는 함수
        :param corpus: 전처리 완료된 말뭉치
        :return: 분리된 청크
        """

        # 청크 사이즈 선택
        chunk_size_number = input("chunk_size를 선택해주세요. 기본값은 1500입니다.\n"
                                  "1: 1500\n"
                                  "2: 2000\n"
                                  "3: 2500\n"
                                  "4: 3000\n"
                                  "5: 3500\n"
                                  "6: 4000\n\n"
                                  "선택 번호: ")

        chunk_size_checker = {
            '1': 1500,
            '2': 2000,
            '3': 2500,
            '4': 3000,
            '5': 3500,
            '6': 4000
        }

        chunk_size = chunk_size_checker.get(chunk_size_number, 1500)

        # 오버랩 사이즈 선택
        overlap_size_number = input("chunk_overlap를 선택해주세요. 기본값은 0입니다.\n"
                                    "1: 0\n"
                                    "2: 100\n"
                                    "3: 200\n"
                                    "4: 300\n"
                                    "5: 400\n"
                                    "6: 500\n\n"
                                    "선택 번호: ")

        overlap_size_checker = {
            '1': 0,
            '2': 100,
            '3': 200,
            '4': 300,
            '5': 400,
            '6': 500
        }

        overlap_size = overlap_size_checker.get(overlap_size_number, 0)

        c_text_splitter = CharacterTextSplitter.from_tiktoken_encoder(
            separator="---",
            chunk_size=chunk_size,
            chunk_overlap=overlap_size,
            # encoding_name="cl100k_base",  # gpt-4
            # encoding_name="o200k_base"  # gpt-4o
        )

        # c_text_splitter = KonlpyTextSplitter.from_tiktoken_encoder(
        #     encoding_name="o200k_base",  # gpt-4o
        #     # encoding_name=""cl100k_base"",  # gpt-4
        #     separator="---",
        #     chunk_size=1500,
        #     chunk_overlap=0
        # )

        text_documents = c_text_splitter.split_documents(corpus)

        return text_documents

    def rc_text_split(self, corpus: List[str]) -> List[str]:
        """
        RecursiveCharacterTextSplitter를 사용하여 문서를 분할하도록 하는 함수
        :param corpus: 전처리 완료된 말뭉치
        :return: 분리된 청크
        """

        # 청크 사이즈 선택
        chunk_size_number = input("chunk_size를 선택해주세요. 기본값은 1500입니다.\n"
                                  "1: 1500\n"
                                  "2: 2000\n"
                                  "3: 2500\n"
                                  "4: 3000\n"
                                  "5: 3500\n"
                                  "6: 4000\n\n"
                                  "선택 번호: ")

        chunk_size_checker = {
            '1': 1500,
            '2': 2000,
            '3': 2500,
            '4': 3000,
            '5': 3500,
            '6': 4000
        }

        chunk_size = chunk_size_checker.get(chunk_size_number, 1500)

        # 오버랩 사이즈 선택
        overlap_size_number = input("chunk_overlap를 선택해주세요. 기본값은 0입니다.\n"
                                    "1: 0\n"
                                    "2: 100\n"
                                    "3: 200\n"
                                    "4: 300\n"
                                    "5: 400\n"
                                    "6: 500\n\n"
                                    "선택 번호: ")

        overlap_size_checker = {
            '1': 0,
            '2': 100,
            '3': 200,
            '4': 300,
            '5': 400,
            '6': 500
        }

        overlap_size = overlap_size_checker.get(overlap_size_number, 0)

        rc_text_splitter = RecursiveCharacterTextSplitter.from_tiktoken_encoder(
            # separators=["---", "\n\n", "\n"],
            separators=["\n\n", "\n", " "],
            chunk_size=chunk_size,
            chunk_overlap=overlap_size,
            model_name="gpt-4o"  # o200k_base
            # model_name="gpt-4"  # cl100k_base
        )

        text_documents = rc_text_splitter.split_documents(corpus)
        print(f"문서가 {len(text_documents)}개로 분할되었습니다.\n")
        print(f"첫 번째 문서의 내용: {text_documents[100].page_content}\n")

        return text_documents, chunk_size, overlap_size

    def embedding_model_select_save(self):
        """
        임베딩 모델을 선택하고 저장하는 함수
        """
        embedding_model_number = input(
            "문서 임베딩에 사용할 임베딩 모델을 고르시오. 고르지 않을 경우 HuggingFaceEmbeddings 모델을 기본으로 사용합니다.\n"
            "1: OpenAIEmbeddings()\n"
            "2: UpstageEmbeddings()\n"
            "3: HuggingFaceEmbeddings()\n\n "
            "선택 번호 : ")

        if embedding_model_number == '1':
            models = {
                '1': "text-embedding-3-small",
                '2': "text-embedding-3-large",
                '3': "text-embedding-ada-002"
            }

            embedding_model_number = input("사용할 OpenAI Embedding Model을 선택하시오. 기본으로 text-embedding-3-small 모델을 사용합니다.\n"
                                           "1: text-embedding-3-small\n"
                                           "2: text-embedding-3-large\n"
                                           "3: text-embedding-ada-002\n\n"
                                           "선택 번호: ")

            model_name = models.get(embedding_model_number, "text-embedding-3-small")

            model = OpenAIEmbeddings(
                openai_api_key=os.getenv("OPENAI_API_KEY"),
                model=model_name
            )
        elif embedding_model_number == '2':
            model_name = "solar-embedding-1-large"
            model = UpstageEmbeddings(
                api_key=os.getenv("UPSTAGE_API_KEY"),
                model=model_name
            )
        else:
            models = {
                '1': "snunlp/KR-SBERT-V40K-klueNLI-augSTS",
                '2': "jhgan/ko-sroberta-multitask",
                '3': "jhgan/ko-sbert-multitask",
                '4': "jhgan/ko-sroberta-nli",
                '5': "jhgan/ko-sbert-nli",
                '6': "jhgan/ko-sroberta-sts",
                '7': "jhgan/ko-sbert-sts",
                '8': "BAAI/bge-m3",
                '9': "sentence-transformers/LaBSE",
                '10': "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2",
                '11': "sentence-transformers/paraphrase-multilingual-mpnet-base-v2"
            }

            embedding_model_number = input(
                "사용할 HuggingFace Embedding Model을 선택하시오. 기본으로 jhgan/ko-sroberta-multitask 모델을 사용합니다.\n"
                "1: snunlp/KR-SBERT-V40K-klueNLI-augSTS\n"
                "2: jhgan/ko-sroberta-multitask\n"
                "3: jhgan/ko-sbert-multitask\n"
                "4: jhgan/ko-sroberta-nli\n"
                "5: jhgan/ko-sbert-nli\n"
                "6: jhgan/ko-sroberta-sts\n"
                "7: jhgan/ko-sbert-sts\n"
                "8: BAAI/bge-m3\n"
                "9: sentence-transformers/LaBSE\n"
                "10: sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2\n"
                "11: sentence-transformers/paraphrase-multilingual-mpnet-base-v2\n\n"
                "선택 번호: ")

            model_name = models.get(embedding_model_number, "jhgan/ko-sroberta-multitask")

            # model_kwargs = {'device': 'cpu'}  # cpu를 사용하기 위해 설정
            model_kwargs = {'device': 'cuda'}  # gpu를 사용하기 위해 설정
            encode_kwargs = {'normalize_embeddings': True}
            model = HuggingFaceEmbeddings(
                model_name=model_name,
                model_kwargs=model_kwargs,
                encode_kwargs=encode_kwargs
            )

        # return model
        return model, model_name

    def document_embedding_basic(self, docs: List[str], model, save_directory: str) -> Tuple:
        """
        Chroma 벡터저장소를 사용하여 문서를 임베딩하고, BM25Retriever의 기본적인 구조를 통해 문서를 키워드 위주의 임베딩을 진행하여 저장하는 함수
        :param model: 임베딩 모델 종류
        :param save_directory: 벡터저장소 저장 경로
        :param docs: 분할된 문서
        :return: 벡터저장소, BM25(기본)저장소
        """

        print("\n잠시만 기다려주세요.\n\n")

        # 벡터저장소가 이미 존재하는지 확인
        if os.path.exists(save_directory):
            shutil.rmtree(save_directory)
            print(f"디렉토리 {save_directory}가 삭제되었습니다.\n")

        print("문서 벡터화를 시작합니다. ")
        db = Chroma.from_documents(docs, model, persist_directory=save_directory)
        bm_db = BM25Retriever.from_documents(
            docs
        )
        print("새로운 Chroma 데이터베이스가 생성되었습니다.\n")

        return db, bm_db

    def document_embedding_kiwi(self, docs: List[str], model, save_directory: str) -> Tuple:
        """
        Chroma 벡터저장소를 사용하여 문서를 임베딩하고, BM25Retriever에 한글 형태소 분석기(Kiki)를 통해 문서를 키워드 위주의 임베딩을 진행하여 저장하는 함수
        :param model: 임베딩 모델 종류
        :param save_directory: 벡터저장소 저장 경로
        :param docs: 분할된 문서
        :return: 벡터저장소, BM25(Kiwi 한글 형태소 분석기)저장소
        """

        print("\n잠시만 기다려주세요.\n\n")

        # 벡터저장소가 이미 존재하는지 확인
        if os.path.exists(save_directory):
            shutil.rmtree(save_directory)
            print(f"디렉토리 {save_directory}가 삭제되었습니다.\n")

        print("문서 벡터화를 시작합니다. ")
        db = Chroma.from_documents(docs, model, persist_directory=save_directory)
        bm_db = BM25Retriever.from_documents(
            docs,
            preprocess_func=self.kiwi_tokenize
        )
        print("새로운 Chroma 데이터베이스가 생성되었습니다.\n")

        return db, bm_db

    def document_embedding_kkma(self, docs: List[str], model, save_directory: str) -> Tuple:
        """
        Chroma 벡터저장소를 사용하여 문서를 임베딩하고, BM25Retriever에 한글 형태소 분석기(Kiki)를 통해 문서를 키워드 위주의 임베딩을 진행하여 저장하는 함수
        :param model: 임베딩 모델 종류
        :param save_directory: 벡터저장소 저장 경로
        :param docs: 분할된 문서
        :return: 벡터저장소, BM25(Kiwi 한글 형태소 분석기)저장소
        """

        print("\n잠시만 기다려주세요.\n\n")

        # 벡터저장소가 이미 존재하는지 확인
        if os.path.exists(save_directory):
            shutil.rmtree(save_directory)
            print(f"디렉토리 {save_directory}가 삭제되었습니다.\n")

        print("문서 벡터화를 시작합니다. ")
        db = Chroma.from_documents(docs, model, persist_directory=save_directory)
        bm_db = BM25Retriever.from_documents(
            docs,
            preprocess_func=self.kkma_tokenize
        )
        print("새로운 Chroma 데이터베이스가 생성되었습니다.\n")

        return db, bm_db

    def document_embedding_okt(self, docs: List[str], model, save_directory: str) -> Tuple:
        """
        Chroma 벡터저장소를 사용하여 문서를 임베딩하고, BM25Retriever에 한글 형태소 분석기(Kiki)를 통해 문서를 키워드 위주의 임베딩을 진행하여 저장하는 함수
        :param model: 임베딩 모델 종류
        :param save_directory: 벡터저장소 저장 경로
        :param docs: 분할된 문서
        :return: 벡터저장소, BM25(Kiwi 한글 형태소 분석기)저장소
        """

        print("\n잠시만 기다려주세요.\n\n")

        # 벡터저장소가 이미 존재하는지 확인
        if os.path.exists(save_directory):
            shutil.rmtree(save_directory)
            print(f"디렉토리 {save_directory}가 삭제되었습니다.\n")

        print("문서 벡터화를 시작합니다. ")
        db = Chroma.from_documents(docs, model, persist_directory=save_directory)
        bm_db = BM25Retriever.from_documents(
            docs,
            preprocess_func=self.okt_tokenize
        )
        print("새로운 Chroma 데이터베이스가 생성되었습니다.\n")

        return db, bm_db

    def chat_llm(self):
        """
        채팅에 사용되는 거대언어모델 생성 함수
        :return: 답변해주는 거대언어모델
        """
        load_dotenv('.env')

        while True:
            model_check = input(
                "채팅에 사용할 모델을 고르시오. 고르지 않을 경우 Google Gemini-1.5 Pro 모델을 기본으로 사용합니다.\n"
                "1: GPT-4o-mini\n2: GPT-4-turbo\n3: GPT-4o\n"
                "4: Claude-3-sonnet\n5: Claude-3-opus\n6: Claude-3.5-sonnet-20240620\n"
                "7: Google Gemini-Pro\n"
                "8: Google Gemma-2-9b-it\n9: Meta Llama-3.1-Instruct\n10: Mistral-Instruct-v0.3\n11: Qwen2.5-7B-instruct\n"
                "12: EEVE Korean\n13: Llama-3-MAAL-8B-Instruct-v0.1\n\n"
                "선택 번호 : ")

            if model_check in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13']:
                break
            else:
                print("잘못된 입력입니다. 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13 중 하나를 선택해주세요.\n")

        model_info = self.get_model_info(model_check)
        model_class = model_info["model_class"]
        model_kwargs = {
            "temperature": 0,
        }

        if model_class == ChatOpenAI:
            model_kwargs.update({
                "model_name": model_info["model_name"],
                "streaming": True,
                "callbacks": [StreamingStdOutCallbackHandler()],
            })
            if "base_url" in model_info:
                model_kwargs["base_url"] = model_info["base_url"]
            if "api_key" in model_info:
                model_kwargs["api_key"] = model_info["api_key"]
        elif model_class == ChatAnthropic:
            model_kwargs.update({
                "model_name": model_info["model_name"],
                "streaming": True,
                "callbacks": [StreamingStdOutCallbackHandler()],
            })
            os.getenv("ANTHROPIC_API_KEY")
        elif model_class == ChatGoogleGenerativeAI:
            os.getenv("GOOGLE_API_KEY")
            model_kwargs.update({
                "model": model_info["model_name"],
            })

        try:
            llm = model_class(**model_kwargs)
        except Exception as e:
            print(f"Error initializing the model: {str(e)}")
            return None, None

        return llm, model_check

    def get_model_info(self, model_check: str) -> dict:
        """
        선택된 모델에 알맞은 정보를 가공하는 함수
        """
        models = {
            "1": {"model_name": "gpt-4o-mini", "model_class": ChatOpenAI},
            "2": {"model_name": "gpt-4-turbo", "model_class": ChatOpenAI},
            "3": {"model_name": "gpt-4o", "model_class": ChatOpenAI},
            "4": {"model_name": "claude-3-sonnet-20240229", "model_class": ChatAnthropic},
            "5": {"model_name": "claude-3-opus-20240229", "model_class": ChatAnthropic},
            "6": {"model_name": "claude-3-5-sonnet-20240620", "model_class": ChatAnthropic},
            "7": {"model_name": "gemini-1.5-pro-exp-0827", "model_class": ChatGoogleGenerativeAI},
            "8": {"model_name": "bartowski/gemma-2-9b-it-GGUF", "model_class": ChatOpenAI,
                  "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
            "9": {"model_name": "lmstudio-community/Meta-Llama-3.1-8B-Instruct-GGUF", "model_class": ChatOpenAI,
                  "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
            "10": {"model_name": "lmstudio-community/Mistral-7B-Instruct-v0.3-GGUF", "model_class": ChatOpenAI,
                   "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
            "11": {"model_name": "lmstudio-community/Qwen2.5-7B-Instruct-GGUF", "model_class": ChatOpenAI,
                   "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
            "12": {"model_name": "teddylee777/EEVE-Korean-Instruct-10.8B-v1.0-gguf", "model_class": ChatOpenAI,
                   "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
            "13": {"model_name": "asiansoul/Llama-3-MAAL-8B-Instruct-v0.1-GGUF", "model_class": ChatOpenAI,
                   "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
        }

        return models.get(model_check)

    def format_docs(self, docs):
        return "\n\n".join(document.page_content for document in docs)

    def kiwi_tokenize(self, text):
        return [token.form for token in self.kiwi.tokenize(text)]

    def kkma_tokenize(self, text):
        return [token for token in self.kkma.morphs(text)]

    def okt_tokenize(self, text):
        return [token for token in self.okt.morphs(text)]

    def reorder_documents(self, docs):
        # 재정렬
        reordering = LongContextReorder()
        reordered_docs = reordering.transform_documents(docs)
        combined = self.format_docs(reordered_docs)

        return combined

    def db_qna_selfQuery(self, llm, db, query):
        """
        셀프 쿼리를 사용하여 문서 검색 후 적절한 답변을 찾아서 답하도록 하는 함수
        :param llm: 거대 언어 모델
        :param db: 벡터스토어
        :param query: 사용자 질문
        """
        metadata_field_info = [
            AttributeInfo(
                name="source",
                description="This document contains information about the College Scholastic Ability Test (CSAT) based regular admission process for Dong-Eui University.",
                type="string",
            ),
        ]

        document_content_description = "This document contains information about the College Scholastic Ability Test (CSAT) based regular admission process for Dong-Eui University."

        self_query = SelfQueryRetriever.from_llm(
            llm,
            db,
            document_content_description,
            metadata_field_info,
            verbose=True,
            enable_limit=True,  # 검색 결과 제한 기능을 활성화합니다.
            search_kwargs={'k': 3},
        )

        db = db.as_retriever(
            search_kwargs={'k': 3},
        )

        # 앙상블 retriever를 초기화합니다.
        ensemble_retriever = EnsembleRetriever(
            retrievers=[self_query, db],
            weights=[0.5, 0.5],
            search_type="mmr",
        )

        prompt = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    You are a specialized AI for question-and-answer tasks.
                    You must answer questions based solely on the Context provided.
                    For questions about predicting successful applicants, base your answers on data from either the initial successful applicants or the final enrolled students.
                    If no Context is provided, you must instruct to inquire at "https://ipsi.deu.ac.kr/main.do".

                    Context: {context}
                    """,
                ),
                ("human", "Question: {question}"),
            ]
        )

        chain = {
                    "context": ensemble_retriever | RunnableLambda(self.reorder_documents),
                    "question": RunnablePassthrough()
                } | prompt | llm | StrOutputParser()

        start_time = time.time()

        response = chain.invoke(query)

        if not isinstance(llm, ChatOpenAI):
            print("\n\n{}".format(response))

        end_time = time.time()

        response_time = (end_time - start_time) * 1000
        print(f"실행 시간: {response_time}")

        return response

    def db_qna_ensemble(self, llm, bm_db, db, query):
        """
        BM25Retriever와 Chroma 벡터스토어를 앙상블하여 문서 검색 후 적절한 답변을 찾아서 답하도록 하는 함수
        :param llm: 거대 언어 모델
        :param bm_db: BM25Retriever
        :param db: 벡터스토어
        :param query: 사용자 질문
        """

        db = db.as_retriever(
            search_kwargs={'k': 3},
        )
        bm_db.k = 3  # BM25Retriever의 검색 결과 개수를 3로 설정

        # 앙상블 retriever를 초기화합니다.
        ensemble_retriever = EnsembleRetriever(
            retrievers=[bm_db, db],
            weights=[0.7, 0.3],
            search_type="mmr",
        )

        """
You are a helpful AI Assistant. Please answer in Korean.
You must answer questions based solely on the Context provided.
If no Context is provided, you must instruct to inquire at "https://ipsi.deu.ac.kr/main.do".
        """

        # 질문에 대한 답변을 찾기 위한 프롬프트
        prompt = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """                    
You are a helpful AI Assistant. Please answer in Korean.

**Context Format:**

You will be provided with Context in a structured format enclosed within `<row>` tags. Each `<row>` contains a `<title>` element representing the title, followed by a series of `<attr>` (attribute) and `<value>` pairs. Attributes can be nested to represent hierarchical relationships.

**Structure Examples:**

- **Simple Attributes:**
```xml
<row> <title>표 제목 1</title>에서 <attr>속성 1</attr>은 <value>값 1</value>이며, <attr>속성 2</attr>은 <value>값 2</value>이다. </row>
```

- Nested Attributes:
```xml
<row><title>표  제목 1</title>에서 <attr>표 속성 1</attr>은 <value>값 1</value>이며, <attr>표 속성 2</attr>은 <value>값 2</value>이며, <attr>표 속성 3</attr>은 <value>값 3</value>이며, <attr>표 속성 4</attr>의 <attr>표 속성 4-1</attr>은 <value>값 4</value>이며, <attr>표 속성 4</attr>의 <attr>표 속성 4-2</attr>은 <value>값 5</value>이며, <attr>표 속성 4</attr>의 <attr>표 속성 4-3</attr>은 <value>값 6</value>이다.</row>
```

```xml
<row><title>표  제목 3</title>에서 <attr>표 속성 1</attr>은 <value>값 1</value>이며, <attr>표 속성 2</attr>은 <value>값 2</value>이며, <attr>표 속성 3</attr>은 <value>값 3</value>이며, <attr>표 속성 4</attr>의 <attr>표 속성 4-1</attr>의 <attr>표 속성 4-1-1</attr>은 <value>값 4</value>이며, <attr>표 속성 4</attr>의 <attr>표 속성 4-1</attr>의 <attr>표 속성 4-1-2</attr>은 <value>값 5</value>이며, <attr>표 속성 4</attr>의 <attr>표 속성 4-2</attr>은 <value>값 6</value>이다.</row>
```

**Your Task:**

- **Extract and Interpret Information:**

  - Carefully parse the Context, paying attention to the hierarchy and relationships between attributes.
  - Accurately extract values corresponding to their attributes.
- Answer Based Solely on the Context:
  - Use only the information explicitly provided in the Context.
  - Do not introduce any external information or assumptions.
- Provide Clear and Concise Answers:
  - Present the information in an organized manner.
  - Ensure your response is easy to understand for the user.
- If No Context Is Provided:
  - Instruct the user to inquire at "https://ipsi.deu.ac.kr/main.do".

**Important Notes:**

- **Language**: All your responses should be in Korean.
- **Accuracy**: Double-check the extracted information for correctness.
- **Clarity**: Use appropriate formatting or bullet points if necessary to enhance readability.

**Context:**

{context}
                    """,
                ),
                ("human", "Question: {question}"),
            ]
        )

        chain = {
                    "context": ensemble_retriever | RunnableLambda(self.reorder_documents),
                    "question": RunnablePassthrough()
                } | prompt | llm | StrOutputParser()

        response = chain.invoke(query)

        if not isinstance(llm, ChatOpenAI):
            print("\n\n{}".format(response))

        return response

    def gpt_score(self, question, correct_answer, answer):
        """
        GPTScore Prompt Template을 사용하여 정답과 답변을 비교하여 점수를 매기는 함수
        :param question: 질문
        :param correct_answer: 모범 응답
        :param answer: 거대언어모델이 생성한 응답
        :return: GPTScore 평가 결과
        """

        # GPT-4o-mini 선언
        llm = ChatOpenAI(
            model="gpt-4o-mini",
            api_key=os.getenv("OPENAI_API_KEY"),
            temperature=0,
            streaming=True,
            callbacks=[StreamingStdOutCallbackHandler()],
        )

        # GPTScore Prompt Template
        gpt_score_prompt = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    # GPTScore Prompt Template

                    ## Factuality
                    Evaluate Factuality in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the factuality of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Factuality (1-5): Does the generated response accurately reflect the factual statements found in the source text? A score of 1 means that the response contains multiple inaccuracies or fabricated information, while a score of 5 means that the response is entirely accurate and preserves all factual details from the source text.

                    **Evaluation Steps**:
                    1. Carefully read the source text to identify key factual statements and details.
                    2. Review the generated response and compare it to the source text, focusing on the accuracy and integrity of the facts presented.
                    3. Assign a score for factuality on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    ## Consistency
                    Evaluate Consistency in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the consistency of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Consistency (1-5): Does the generated response consistently provide information that aligns with the source text? A score of 1 means that the response contains contradictory or conflicting information, while a score of 5 means that the response is fully consistent with no discrepancies.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand the key points and details.
                    2. Review the generated response and compare it to the source text, focusing on the consistency of the information provided.
                    3. Provide a detailed explanation of your evaluation, noting any inconsistencies or confirming the consistency of the response.
                    4. Assign a score for consistency on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the consistency, highlighting specific aspects where the generated response aligns or diverges from the source text.

                    ## Relevance
                    Evaluate Relevance in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the relevance of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Relevance (1-5): How well does the generated response relate to the source text? A score of 1 means that the response is largely irrelevant or off-topic, while a score of 5 means that the response is highly relevant and directly addresses the key points of the source text.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its main topics and key points.
                    2. Review the generated response and compare it to the source text, focusing on how well it addresses the main topics and key points.
                    3. Provide a detailed explanation of your evaluation, highlighting specific areas where the generated response is relevant or irrelevant to the source text.
                    4. Assign a score for relevance on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the relevance, highlighting specific aspects where the generated response aligns or diverges from the main topics of the source text.

                    ## Fluency
                    Evaluate Fluency in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the fluency of the generated response by assessing its grammatical correctness and readability.

                    **Evaluation Criteria**:
                    - Fluency (1-5): How well is the generated response written in terms of grammar, syntax, and overall readability? A score of 1 means the response is poorly written, with numerous grammatical errors or awkward phrasing, while a score of 5 means the response is highly fluent, with excellent grammar and smooth readability.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its style and tone.
                    2. Review the generated response, focusing on its grammatical correctness, sentence structure, and overall readability.
                    3. Provide a detailed explanation of your evaluation, noting any grammatical errors, awkward phrasing, or confirming the fluency of the response.
                    4. Assign a score for fluency on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the fluency, highlighting specific aspects where the generated response is grammatically correct, easy to read, or where it may have issues with fluency.

                    ## Coherence
                    Evaluate Coherence in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the coherence of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Coherence (1-5): How well does the generated response make sense as a unified piece of text? A score of 1 means that the response is disjointed or lacks logical flow, while a score of 5 means that the response is well-structured and logically consistent throughout.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its overall structure and key points.
                    2. Review the generated response and assess its logical flow, structure, and overall coherence.
                    3. Provide a detailed explanation of your evaluation, noting any logical inconsistencies or confirming the coherence of the response.
                    4. Assign a score for coherence on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the coherence, highlighting specific aspects where the generated response aligns or diverges in its logical flow and structure compared to the source text.

                    ## Accuracy
                    Evaluate Accuracy in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the accuracy of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Accuracy (1-5): Are there any inaccuracies, omissions, or unfactual content in the generated response? A score of 1 means that the response contains significant inaccuracies or incorrect information, while a score of 5 means that the response is fully accurate and correctly reflects the source text.

                    **Evaluation Steps**:
                    1. Carefully read the source text to identify all the key facts and details.
                    2. Review the generated response and compare it to the source text, focusing on the accuracy of the information provided.
                    3. Provide a detailed explanation of your evaluation, highlighting any inaccuracies, omissions, or confirming the accuracy of the response.
                    4. Assign a score for accuracy on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the accuracy, highlighting specific aspects where the generated response aligns with or diverges from the facts in the source text.

                    ## Multidimensional Quality
                    Evaluate Multidimensional Quality in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the overall quality of the generated response across multiple dimensions, including factuality, coherence, relevance, and accuracy.

                    **Evaluation Criteria**:
                    - Multidimensional Quality (1-5): How well does the generated response perform across all relevant quality dimensions (factuality, coherence, relevance, accuracy)? A score of 1 means the response performs poorly in most or all dimensions, while a score of 5 means the response performs excellently across all dimensions.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand the main content and quality dimensions.
                    2. Review the generated response, assessing its performance in terms of factuality, coherence, relevance, and accuracy.
                    3. Provide a detailed explanation of your evaluation, highlighting the strengths and weaknesses of the response across all quality dimensions.
                    4. Assign a score for overall multidimensional quality on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the overall quality, considering the factuality, coherence, relevance, and accuracy of the generated response.

                    ## Semantic Appropriateness
                    Evaluate Semantic Appropriateness in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the semantic appropriateness of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Semantic Appropriateness (1-5): How well does the generated response convey meaning that is appropriate and aligned with the context of the source text? A score of 1 means the response is semantically inappropriate or off-context, while a score of 5 means the response is fully appropriate and semantically consistent with the source text.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its meaning and context.
                    2. Review the generated response and assess whether the meaning it conveys is appropriate and aligned with the source text.
                    3. Provide a detailed explanation of your evaluation, highlighting areas where the generated response is semantically appropriate or where it diverges from the intended meaning.
                    4. Assign a score for semantic appropriateness on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the semantic appropriateness, highlighting specific aspects where the generated response aligns or diverges from the meaning and context of the source text.

                    ## Understandability
                    Evaluate Understandability in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the understandability of the generated response by considering how clear and easy it is to comprehend.

                    **Evaluation Criteria**:
                    - Understandability (1-5): How easy is it to understand the generated response? A score of 1 means that the response is confusing, unclear, or difficult to comprehend, while a score of 5 means that the response is very clear, straightforward, and easy to understand.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its main points and context.
                    2. Review the generated response, focusing on how clearly the information is presented and whether it is easy to follow.
                    3. Provide a detailed explanation of your evaluation, noting any areas where the response is particularly clear or where it may be confusing.
                    4. Assign a score for understandability on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the understandability, highlighting specific aspects where the generated response is clear or where it may be difficult to comprehend.

                    ## Output Score
                    **Example**:
                    Question:
                    {question}
                    
                    Source Text:
                    {source_text}

                    Generated Response:
                    {generated_response}

                    1. Factuality Score (1-5):
                    2. Consistency Score (1-5):
                    3. Relevance Score (1-5):
                    4. Fluency Score (1-5):
                    5. Coherence Score (1-5):
                    6. Accuracy Score (1-5):
                    7. Multidimensional Quality Score (1-5):
                    8. Semantic Appropriateness Score (1-5):
                    9. Understandability Score (1-5):

                    {{
                        "Factuality Score": 4,
                        "Consistency Score": 3,
                        "Relevance Score": 5,
                        "Fluency Score": 4,
                        "Coherence Score": 4,
                        "Accuracy Score": 4,
                        "Multidimensional Quality Score": 4,
                        "Semantic Appropriateness Score": 4,
                        "Understandability Score": 4
                    }}

                    """,
                ),
                ("human", "generated_response: {generated_response}"),
            ]
        )

        gpt_score_chain = {
                              "question": lambda x: question,
                              "source_text": lambda x: correct_answer,
                              "generated_response": RunnablePassthrough()
                          } | gpt_score_prompt | llm | StrOutputParser()

        gpt_score_response = gpt_score_chain.invoke({"source_text": correct_answer, "generated_response": answer})

        if not isinstance(llm, ChatOpenAI):
            print("\n\n{}".format(gpt_score_response))

        return gpt_score_response

    def db_qna(self, llm, db, query, ):
        """
        벡터저장소에서 질문을 검색해서 적절한 답변을 찾아서 답하도록 하는 함수
        :param llm: 거대 언어 모델
        :param db: 벡터스토어
        :param query: 사용자 질문
        :return: 거대언어모델(LLM) 응답 결과
        """

        db = db.as_retriever(
            # search_kwargs={'k': 3},
            search_type="mmr",
            search_kwargs={'k': 3, 'fetch_k': 10},
        )

        prompt = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    You are a specialized AI for question-and-answer tasks.
                    You must answer questions based solely on the Context provided.
                    If no Context is provided, you must instruct to inquire at "https://ipsi.deu.ac.kr/main.do".

                    Context: {context}
                    """,
                ),
                ("human", "Question: {question}"),
            ]
        )

        chain = {
                    "context": db | RunnableLambda(self.format_docs),
                    "question": RunnablePassthrough()
                } | prompt | llm | StrOutputParser()

        response = chain.invoke(query)

        if not isinstance(llm, ChatOpenAI):
            print("\n\n{}".format(response))

        return response


class ExperimentAutomation:
    def __init__(self):
        self.system = "ExperimentAutomation"
        self.chatbot = ChatBotSystem()

    def cosine_similarity(self, a, b):
        """
        코사인 유사도를 확인하기 위한 함수
        :param a: 벡터 a
        :param b: 벡터 b
        """
        dot_product = np.dot(a, b)
        norm_a = np.linalg.norm(a)
        norm_b = np.linalg.norm(b)
        return dot_product / (norm_a * norm_b)

    def save_qna_list_v2(self, q, a, model_answer, model_checker, similarity, embedding_model_name, chunk_size,
                         overlap_size, gpt_score_response):
        """
        질의 응답을 엑셀 파일에 추가하는 함수 (중복 질문 제거)
        @param q: 질의
        @param a: 모범 응답
        @param model_answer: 질의에 대한 거대언어모델 응답
        @param model_checker: 선택한 거대언어모델 이름을 알기 위한 번호
        @param similarity: 질의에 대한 모범 응답과 거대언어모델 응답의 유사도
        @param embedding_model_name: 사용한 임베딩 모델 이름
        @param chunk_size: 문서 분할기의 청크 사이즈
        @param overlap_size: 문서 분할기의 오버랩 사이즈
        """

        embedding_model_name = embedding_model_name.replace('/', '_')
        filename = f'research_result/{embedding_model_name}_({chunk_size}_{overlap_size})_RecursiveCharacterTextSplitter.xlsx'

        filename = f'research_result/v2_({chunk_size}_{overlap_size}).xlsx'

        # model_checker 값을 모델 이름으로 변환
        model_name = ''
        if model_checker == '1':
            model_name = 'GPT-4o-mini'
        elif model_checker == '2':
            model_name = 'GPT-4'
        elif model_checker == '3':
            model_name = 'GPT-4o'
        elif model_checker == '4':
            model_name = 'Claude-3-sonnet-20240229'
        elif model_checker == '5':
            model_name = 'Claude-3-opus-20240229'
        elif model_checker == '6':
            model_name = 'Claude-3-5-sonnet-20240620'
        elif model_checker == '7':
            model_name = 'Google Gemini-Pro-exp-0827'
        elif model_checker == '8':
            model_name = 'Google Gemma-2-9b-it'
        elif model_checker == '9':
            model_name = 'Meta Llama-3.1-Instruct'
        elif model_checker == '10':
            model_name = 'Mistral-7B-Instruct-v0.3'
        elif model_checker == '11':
            model_name = 'Qwen2.5-7B-instruct'
        elif model_checker == '12':
            model_name = 'EEVE-Korean-Instruct-10.8B-v1.0'
        elif model_checker == '13':
            model_name = 'Llama-3-MAAL-8B-Instruct-v0.1'
        try:
            # 기존 엑셀 파일 열기
            workbook = load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            # 파일이 없는 경우 새로운 엑셀 파일 생성
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['A1'] = '거대언어모델'
            sheet['B1'] = 'GPTScore'
            sheet['C1'] = '유사도'
            sheet['D1'] = '질의'
            sheet['E1'] = '응답'
            sheet['F1'] = '모범 응답'

        # '모범 응답' 헤더 추가 (파일이 이미 존재하는 경우에도)
        if not isinstance(sheet['F1'].value, str) or sheet['F1'].value != '모범 응답':
            sheet['F1'] = '모범 응답'

        # 기존 질문 목록 가져오기
        existing_questions = set((cell.value, sheet.cell(row=cell.row, column=1).value) for cell in sheet['D'][1:])

        # 중복 질문 확인
        if (q, model_name) not in existing_questions:
            # 새로운 행에 데이터 추가
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1, value=model_name)
            sheet.cell(row=row, column=2, value=gpt_score_response)
            sheet.cell(row=row, column=3, value=similarity)
            sheet.cell(row=row, column=4, value=q)
            sheet.cell(row=row, column=5, value=a)
            sheet.cell(row=row, column=6, value=model_answer)

        # 거대언어모델별로 정렬
        # data = list(sheet.values)[1:]
        # # print(f"data: {data}")
        # data.sort(key=lambda x: (x[0], x[1]))
        #
        # # 정렬된 데이터로 시트 업데이트
        # sheet.delete_rows(2, sheet.max_row)
        # for row, row_data in enumerate(data, start=2):
        #     for col, value in enumerate(row_data, start=1):
        #         sheet.cell(row=row, column=col, value=value)

        # 엑셀 파일 저장
        workbook.save(filename)

        if model_name == 'Google Gemini-Pro':
            time.sleep(20)

    def auto_question(self, llm, db, bm_db, model_num, embedding_model, embedding_model_name, chunk_size, overlap_size):
        """
        질문 리스트를 기반으로 자동으로 질문하고 답변을 받아 엑셀 파일에 저장하는 함수
        :param llm: 거대언어모델 종류
        :param db: 기본 벡터저장소
        :param bm_db: bm 벡터저장소
        :param model_num: 선택한 거대언어모델 이름을 위한 번호
        :param embedding_model: 임베딩 모델
        :param embedding_model_name: 선택한 임베딩 모델 이름        
        :param chunk_size: 문서 분할기의 청크 사이즈        
        :param overlap_size: 문서 분할기의 오버랩 사이즈
        """

        df = pd.read_excel("test_automation/qna.xlsx")

        questions_list = df['질의'].tolist()
        correct_answers = df['모범 응답'].tolist()

        for question, correct_answer in zip(questions_list, correct_answers):
            # response = self.chatbot.db_qna(llm, db, question)  # 기본 검색기
            # 거대언어모델이 생성한 응답
            response = self.chatbot.db_qna_ensemble(llm, bm_db, db,
                                                    question)  # 앙상블 검색기 (키워드 기반 문서 검색, 의미적 유사성 기반 문서 검색)
            # response = self.chatbot.db_qna_selfQuery(llm, db, question)  # 앙상블 검색기 (셀프 쿼리 기반 문서 검색, 의미적 유사성 기반 문서 검색,)

            # 코사인 유사도 확인
            temp_model_answer = embedding_model.embed_query(correct_answer)
            temp_response = embedding_model.embed_query(response)
            similarity = self.cosine_similarity(temp_model_answer, temp_response)
            print(f"similarity: {similarity}")

            # GPTScore 평가
            gpt_score_response = self.chatbot.gpt_score(question, correct_answer, response)

            # 파일 저장
            # save_qna_list(question, response, model_num, similarity)
            self.save_qna_list_v2(question, response, correct_answer, model_num, similarity, embedding_model_name,
                                  chunk_size, overlap_size, gpt_score_response)

    def manual_question(self, llm, db, bm_db, model_num, embedding_model):
        check = 'Y'  # 0이면 질문 가능
        while check == 'Y' or check == 'y':
            query = input("질문을 입력하세요 : ")
            # response = self.chatbot.db_qna(llm, db, query)  # 기본 검색기
            response = self.chatbot.db_qna_ensemble(llm, bm_db, db, query)  # 앙상블 검색기 (키워드 기반 문서 검색, 의미적 유사성 기반 문서 검색)
            # response = self.chatbot.db_qna_selfQuery(llm, db, query)  # 앙상블 검색기 (셀프 쿼리 기반 문서 검색, 의미적 유사성 기반 문서 검색,)

            # 코사인 유사도 확인
            # temp_q = embedding_model.embed_query(query)
            # temp_a = embedding_model.embed_query(response)
            # similarity = cosine_similarity(temp_q, temp_a)

            check = input("\n\nY: 계속 질문한다.\nN: 프로그램 종료\n입력: ")

    def score_calculate(self):
        # research_result 폴더 경로
        folder_path = "research_result"
        output_path = "research_result/output"

        # output 폴더가 존재하지 않으면 생성
        os.makedirs(output_path, exist_ok=True)

        # research_result 폴더 내 모든 파일 처리
        for file_name in os.listdir(folder_path):

            if file_name.startswith('~$'):  # 엑셀 임시 파일 오류 방지
                continue

            file_path = os.path.join(folder_path, file_name)

            print(f"file_path: {file_path}")

            # 엑셀 파일만 처리 (파일 확장자가 .xlsx인 경우)
            if file_name.endswith('.xlsx'):
                print("steste")
                try:
                    # Json 전처리
                    score_preprocessing = GPTScorePreprocessing(
                        file_path,  # 각 파일 경로
                        "Sheet",  # 엑셀 파일의 시트명
                        0,  # model_name 열 순번
                        1  # GPTScore 열 순번
                    )

                    # score_preprocessing.load_excel()  # 엑셀 로드
                    # score_preprocessing.extract_json()  # Json 추출
                    # score_preprocessing.normalize_json() # 정규화
                    score_preprocessing.run()  # 엑셀 로드 -> Json 추출 -> 정규화

                    # 파일 이름에서 확장자 제거하여 저장 이름 설정
                    save_name = os.path.splitext(file_name)[0]
                    # save_file_path = os.path.join(output_path, f"{save_name}.json")

                    # 파일이 이미 존재하는지 확인하고 덮어쓰기 처리
                    # if os.path.exists(save_file_path):
                    #     print(f"파일이 이미 존재합니다. 덮어쓰기 합니다: {save_file_path}")

                    score_preprocessing.save_json(
                        save_path=output_path,
                        save_name=save_name  # 파일 이름 기반으로 저장
                    )

                    print(f"Json 전처리 성공: {file_name}")

                except FileNotFoundError as e:
                    print(f"파일을 찾을 수 없습니다: {file_name} - {e}")
                except ValueError as e:
                    print(f"값 오류 발생: {file_name} - {e}")
                except Exception as e:
                    print(f"오류 발생...{file_name}: {e}")


def run():
    """
    챗봇 시작
    Document Load -> Text Splitter -> Ducument Embedding -> VectorStore save -> QA
    """

    chatbot = ChatBotSystem()
    experiment = ExperimentAutomation()

    # 환경변수 로드
    chatbot.load_env()

    # 문서 업로드
    loader = chatbot.docs_load()

    # 문서 분할
    # chunk = chatbot.c_text_split(loader)
    chunk, chunk_size, overlap_size = chatbot.rc_text_split(loader)

    print(f"chunk: {len(chunk)}")

    # 문서 임베딩 및 벡터스토어 저장
    embedding_model, embedding_model_name = chatbot.embedding_model_select_save()
    # print(f"embedding_model_name: {embedding_model_name}")

    # bm25 + 한글 형태소 분석기(kiwi, Kkma, Okt) 추가
    # db, bm_db = chatbot.document_embedding_basic(chunk, embedding_model, save_directory="./chroma_db")
    db, bm_db = chatbot.document_embedding_kiwi(chunk, embedding_model, save_directory="./chroma_db")
    # db, bm_db = chatbot.document_embedding_kkma(chunk, embedding_model, save_directory="./chroma_db")
    # db, bm_db = chatbot.document_embedding_okt(chunk, embedding_model, save_directory="./chroma_db")

    # 채팅에 사용할 거대언어모델(LLM) 선택
    llm, model_num = chatbot.chat_llm()

    # 정보 검색
    q_way = input("1. 질의 수동\n2. 질의 자동(실험용)\n\n사용할 방식을 선택하시오(기본값 수동): ")

    if q_way == '2':
        experiment.auto_question(llm, db, bm_db, model_num, embedding_model, embedding_model_name, chunk_size,
                                 overlap_size)
    else:
        experiment.manual_question(llm, db, bm_db, model_num, embedding_model)


if __name__ == "__main__":
    run()

    experiment = ExperimentAutomation()
    experiment.score_calculate()
