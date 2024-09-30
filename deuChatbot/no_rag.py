from langchain_community.document_loaders import PyPDFLoader
from langchain_chroma import Chroma

import os
from dotenv import load_dotenv
import openpyxl
from openpyxl import load_workbook
import numpy as np
import pandas as pd

from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.runnables import RunnablePassthrough, RunnableLambda
from langchain_huggingface import HuggingFaceEmbeddings

from langchain_openai import ChatOpenAI
from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler

from implements.preprocessing import GPTScorePreprocessing


class NoRAG:
    def __init__(self):
        self.system = "system"

    def documentLoader(self):
        # Load a PDF document
        docs = PyPDFLoader("corpus/2024학년도 정시 모집요강(원본 배포용 제거).pdf").load()

        print(docs)

        return docs

    def embeddingModel(self):
        model_kwargs = {'device': 'cuda'}  # gpu를 사용하기 위해 설정
        encode_kwargs = {'normalize_embeddings': True}
        model = HuggingFaceEmbeddings(
            model_name="BAAI/bge-m3",
            model_kwargs=model_kwargs,
            encode_kwargs=encode_kwargs
        )

        return model

    def vectorStore(self, embedding_model, docs):
        save_directory = "no_rag_vectorstore"
        # 벡터저장소가 이미 존재하는지 확인
        if os.path.exists(save_directory):
            # shutil.rmtree(save_directory)
            # print(f"디렉토리 {save_directory}가 삭제되었습니다.\n")
            db = Chroma(persist_directory=save_directory, embedding_function=embedding_model)
            retriever = db.as_retriever(
                search_kwargs={'k': 1},
            )
            print("기존 Chroma 데이터베이스를 불러왔습니다.\n")

            return retriever

        print("문서 벡터화를 시작합니다. ")
        db = Chroma.from_documents(docs, embedding_model, persist_directory=save_directory)
        retriever = db.as_retriever(
            search_kwargs={'k': 1},
        )
        print("새로운 Chroma 데이터베이스가 생성되었습니다.\n")

        return retriever

    def qna(self, retriever):
        load_dotenv('.env')
        os.getenv("LANGCHAIN_TRACING_V2")
        os.getenv("LANGCHAIN_ENDPOINT")
        os.getenv("LANGCHAIN_API_KEY")
        os.getenv("OPENAI_API_KEY")
        os.getenv("LM_URL")
        os.getenv("LM_LOCAL_URL")

        # OpenAI
        # llm = ChatOpenAI(
        #     model_name="gpt-4o-mini",
        #     api_key=os.getenv("OPENAI_API_KEY"),
        #     temperature=0,
        #     streaming=True,
        #     callbacks=[StreamingStdOutCallbackHandler()],
        # )

        # LM Studio
        llm = ChatOpenAI(
            # model_name="lmstudio-community/Mistral-7B-Instruct-v0.3-GGUF",
            # model_name="lmstudio-community/Meta-Llama-3.1-8B-Instruct-GGUF",
            # model_name="bartowski/gemma-2-9b-it-GGUF",
            model_name="lmstudio-community/Qwen2.5-7B-Instruct-GGUF",
            base_url=os.getenv("LM_URL"),
            api_key="lm-studio",
            temperature=0,
            streaming=True,
            callbacks=[StreamingStdOutCallbackHandler()],
        )

        prompt = ChatPromptTemplate.from_messages(
            [
                ("system", "You are a helpful AI Assistant. Please answer in Korean. Context: {context}"),
                ("human", "Question: {question}"),
            ]
        )

        chain = {
                    "context": retriever | RunnableLambda(lambda x: x),
                    "question": RunnablePassthrough(),
                } | prompt | llm | StrOutputParser()

        return chain

    def qna_init(self):
        docs = self.documentLoader()
        embedding_model = self.embeddingModel()
        retriever = self.vectorStore(embedding_model, docs)
        chain = self.qna(retriever)

        return embedding_model, chain


class Evaluation:
    def __init__(self):
        self.system = "Evaluation"

    def cosine_similarity(self, a, b):
        """
        코사인 유사도를 확인하기 위한 함수
        :param a: 벡터 a
        :param b: 벡터 b
        """
        dot_product = np.dot(a, b)
        norm_a = np.linalg.norm(a)
        norm_b = np.linalg.norm(b)
        return dot_product / (norm_a * norm_b)

    def gpt_score(self, question, correct_answer, answer):
        """
        GPTScore Prompt Template을 사용하여 정답과 답변을 비교하여 점수를 매기는 함수
        :param question: 질문
        :param correct_answer: 모범 응답
        :param answer: 거대언어모델이 생성한 응답
        :return: GPTScore 평가 결과
        """

        # GPT-4o-mini 선언
        llm = ChatOpenAI(
            model="gpt-4o-mini",
            api_key=os.getenv("OPENAI_API_KEY"),
            temperature=0,
            streaming=True,
            callbacks=[StreamingStdOutCallbackHandler()],
        )

        # GPTScore Prompt Template
        gpt_score_prompt = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    # GPTScore Prompt Template

                    ## Factuality
                    Evaluate Factuality in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the factuality of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Factuality (1-5): Does the generated response accurately reflect the factual statements found in the source text? A score of 1 means that the response contains multiple inaccuracies or fabricated information, while a score of 5 means that the response is entirely accurate and preserves all factual details from the source text.

                    **Evaluation Steps**:
                    1. Carefully read the source text to identify key factual statements and details.
                    2. Review the generated response and compare it to the source text, focusing on the accuracy and integrity of the facts presented.
                    3. Assign a score for factuality on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    ## Consistency
                    Evaluate Consistency in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the consistency of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Consistency (1-5): Does the generated response consistently provide information that aligns with the source text? A score of 1 means that the response contains contradictory or conflicting information, while a score of 5 means that the response is fully consistent with no discrepancies.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand the key points and details.
                    2. Review the generated response and compare it to the source text, focusing on the consistency of the information provided.
                    3. Provide a detailed explanation of your evaluation, noting any inconsistencies or confirming the consistency of the response.
                    4. Assign a score for consistency on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the consistency, highlighting specific aspects where the generated response aligns or diverges from the source text.

                    ## Relevance
                    Evaluate Relevance in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the relevance of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Relevance (1-5): How well does the generated response relate to the source text? A score of 1 means that the response is largely irrelevant or off-topic, while a score of 5 means that the response is highly relevant and directly addresses the key points of the source text.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its main topics and key points.
                    2. Review the generated response and compare it to the source text, focusing on how well it addresses the main topics and key points.
                    3. Provide a detailed explanation of your evaluation, highlighting specific areas where the generated response is relevant or irrelevant to the source text.
                    4. Assign a score for relevance on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the relevance, highlighting specific aspects where the generated response aligns or diverges from the main topics of the source text.

                    ## Fluency
                    Evaluate Fluency in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the fluency of the generated response by assessing its grammatical correctness and readability.

                    **Evaluation Criteria**:
                    - Fluency (1-5): How well is the generated response written in terms of grammar, syntax, and overall readability? A score of 1 means the response is poorly written, with numerous grammatical errors or awkward phrasing, while a score of 5 means the response is highly fluent, with excellent grammar and smooth readability.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its style and tone.
                    2. Review the generated response, focusing on its grammatical correctness, sentence structure, and overall readability.
                    3. Provide a detailed explanation of your evaluation, noting any grammatical errors, awkward phrasing, or confirming the fluency of the response.
                    4. Assign a score for fluency on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the fluency, highlighting specific aspects where the generated response is grammatically correct, easy to read, or where it may have issues with fluency.

                    ## Coherence
                    Evaluate Coherence in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the coherence of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Coherence (1-5): How well does the generated response make sense as a unified piece of text? A score of 1 means that the response is disjointed or lacks logical flow, while a score of 5 means that the response is well-structured and logically consistent throughout.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its overall structure and key points.
                    2. Review the generated response and assess its logical flow, structure, and overall coherence.
                    3. Provide a detailed explanation of your evaluation, noting any logical inconsistencies or confirming the coherence of the response.
                    4. Assign a score for coherence on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the coherence, highlighting specific aspects where the generated response aligns or diverges in its logical flow and structure compared to the source text.

                    ## Accuracy
                    Evaluate Accuracy in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the accuracy of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Accuracy (1-5): Are there any inaccuracies, omissions, or unfactual content in the generated response? A score of 1 means that the response contains significant inaccuracies or incorrect information, while a score of 5 means that the response is fully accurate and correctly reflects the source text.

                    **Evaluation Steps**:
                    1. Carefully read the source text to identify all the key facts and details.
                    2. Review the generated response and compare it to the source text, focusing on the accuracy of the information provided.
                    3. Provide a detailed explanation of your evaluation, highlighting any inaccuracies, omissions, or confirming the accuracy of the response.
                    4. Assign a score for accuracy on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the accuracy, highlighting specific aspects where the generated response aligns with or diverges from the facts in the source text.

                    ## Multidimensional Quality
                    Evaluate Multidimensional Quality in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the overall quality of the generated response across multiple dimensions, including factuality, coherence, relevance, and accuracy.

                    **Evaluation Criteria**:
                    - Multidimensional Quality (1-5): How well does the generated response perform across all relevant quality dimensions (factuality, coherence, relevance, accuracy)? A score of 1 means the response performs poorly in most or all dimensions, while a score of 5 means the response performs excellently across all dimensions.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand the main content and quality dimensions.
                    2. Review the generated response, assessing its performance in terms of factuality, coherence, relevance, and accuracy.
                    3. Provide a detailed explanation of your evaluation, highlighting the strengths and weaknesses of the response across all quality dimensions.
                    4. Assign a score for overall multidimensional quality on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the overall quality, considering the factuality, coherence, relevance, and accuracy of the generated response.

                    ## Semantic Appropriateness
                    Evaluate Semantic Appropriateness in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the semantic appropriateness of the generated response by comparing it to the source text.

                    **Evaluation Criteria**:
                    - Semantic Appropriateness (1-5): How well does the generated response convey meaning that is appropriate and aligned with the context of the source text? A score of 1 means the response is semantically inappropriate or off-context, while a score of 5 means the response is fully appropriate and semantically consistent with the source text.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its meaning and context.
                    2. Review the generated response and assess whether the meaning it conveys is appropriate and aligned with the source text.
                    3. Provide a detailed explanation of your evaluation, highlighting areas where the generated response is semantically appropriate or where it diverges from the intended meaning.
                    4. Assign a score for semantic appropriateness on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the semantic appropriateness, highlighting specific aspects where the generated response aligns or diverges from the meaning and context of the source text.

                    ## Understandability
                    Evaluate Understandability in the Generated Response

                    You will be given a source text and a generated response. Your task is to evaluate the understandability of the generated response by considering how clear and easy it is to comprehend.

                    **Evaluation Criteria**:
                    - Understandability (1-5): How easy is it to understand the generated response? A score of 1 means that the response is confusing, unclear, or difficult to comprehend, while a score of 5 means that the response is very clear, straightforward, and easy to understand.

                    **Evaluation Steps**:
                    1. Carefully read the source text to understand its main points and context.
                    2. Review the generated response, focusing on how clearly the information is presented and whether it is easy to follow.
                    3. Provide a detailed explanation of your evaluation, noting any areas where the response is particularly clear or where it may be confusing.
                    4. Assign a score for understandability on a scale of 1 to 5 based on the Evaluation Criteria.

                    **Important**: Ensure that all responses, including explanations and scores, are generated in Korean.

                    Evaluation Explanation:
                    - Provide an analysis of the understandability, highlighting specific aspects where the generated response is clear or where it may be difficult to comprehend.

                    ## Output Score
                    **Example**:
                    Question:
                    {question}

                    Source Text:
                    {source_text}

                    Generated Response:
                    {generated_response}

                    1. Factuality Score (1-5):
                    2. Consistency Score (1-5):
                    3. Relevance Score (1-5):
                    4. Fluency Score (1-5):
                    5. Coherence Score (1-5):
                    6. Accuracy Score (1-5):
                    7. Multidimensional Quality Score (1-5):
                    8. Semantic Appropriateness Score (1-5):
                    9. Understandability Score (1-5):

                    {{
                        "Factuality Score": 4,
                        "Consistency Score": 3,
                        "Relevance Score": 5,
                        "Fluency Score": 4,
                        "Coherence Score": 4,
                        "Accuracy Score": 4,
                        "Multidimensional Quality Score": 4,
                        "Semantic Appropriateness Score": 4,
                        "Understandability Score": 4
                    }}

                    """,
                ),
                ("human", "generated_response: {generated_response}"),
            ]
        )

        gpt_score_chain = {
                              "question": lambda x: question,
                              "source_text": lambda x: correct_answer,
                              "generated_response": RunnablePassthrough()
                          } | gpt_score_prompt | llm | StrOutputParser()

        gpt_score_response = gpt_score_chain.invoke({"source_text": correct_answer, "generated_response": answer})

        if not isinstance(llm, ChatOpenAI):
            print("\n\n{}".format(gpt_score_response))

        return gpt_score_response


class ExperimentAutomation:
    def __init__(self, embedding_model, chain):
        self.system = "ExperimentAutomation"
        self.evaluate = Evaluation()
        self.embedding_model = embedding_model
        self.chain = chain

    def save_qna_list(self, q, a, model_answer, similarity, gpt_score_response):
        """
        질의 응답을 엑셀 파일에 추가하는 함수 (중복 질문 제거)
        @param q: 질의
        @param a: 모범 응답
        @param model_answer: 질의에 대한 거대언어모델 응답
        @param similarity: 질의에 대한 모범 응답과 거대언어모델 응답의 유사도
        :param gpt_score_response:
        """

        filename = f'research_result/no_rag.xlsx'

        # model_checker 값을 모델 이름으로 변환
        # model_name = 'GPT-4o-mini'
        # model_name = 'Mistral-7B-Instruct-v0.3'
        # model_name = 'Meta-Llama-3.1-8B-Instruct'
        # model_name = 'Gemma-2-9b-it'
        model_name = 'Qwen2.5-7B-instruct'

        # if model_checker == '1':
        #     model_name = 'GPT-4o-mini'
        # elif model_checker == '2':
        #     model_name = 'GPT-4'
        # elif model_checker == '3':
        #     model_name = 'GPT-4o'
        # elif model_checker == '4':
        #     model_name = 'Claude-3-sonnet-20240229'
        # elif model_checker == '5':
        #     model_name = 'Claude-3-opus-20240229'
        # elif model_checker == '6':
        #     model_name = 'Claude-3-5-sonnet-20240620'
        # elif model_checker == '7':
        #     model_name = 'Google Gemini-Pro-exp-0827'
        # elif model_checker == '8':
        #     model_name = 'Google Gemma-2-9b-it'
        # elif model_checker == '9':
        #     model_name = 'Meta Llama-3.1-Instruct'
        # elif model_checker == '10':
        #     model_name = 'Mistral-7B-Instruct-v0.3'
        # elif model_checker == '11':
        #     model_name = 'Qwen2.5-7B-instruct'
        # elif model_checker == '12':
        #     model_name = 'EEVE-Korean-Instruct-10.8B-v1.0'
        # elif model_checker == '13':
        #     model_name = 'Llama-3-MAAL-8B-Instruct-v0.1'

        try:
            # 기존 엑셀 파일 열기
            workbook = load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            # 파일이 없는 경우 새로운 엑셀 파일 생성
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['A1'] = '거대언어모델'
            sheet['B1'] = 'GPTScore'
            sheet['C1'] = '유사도'
            sheet['D1'] = '질의'
            sheet['E1'] = '응답'
            sheet['F1'] = '모범 응답'

        # '모범 응답' 헤더 추가 (파일이 이미 존재하는 경우에도)
        if not isinstance(sheet['F1'].value, str) or sheet['F1'].value != '모범 응답':
            sheet['F1'] = '모범 응답'

        # 기존 질문 목록 가져오기
        existing_questions = set((cell.value, sheet.cell(row=cell.row, column=1).value) for cell in sheet['D'][1:])

        # 중복 질문 확인
        if (q, model_name) not in existing_questions:
            # 새로운 행에 데이터 추가
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1, value=model_name)
            sheet.cell(row=row, column=2, value=gpt_score_response)
            sheet.cell(row=row, column=3, value=similarity)
            sheet.cell(row=row, column=4, value=q)
            sheet.cell(row=row, column=5, value=a)
            sheet.cell(row=row, column=6, value=model_answer)

        # 엑셀 파일 저장
        workbook.save(filename)

        if model_name == 'Google Gemini-Pro':
            import time
            time.sleep(20)

    def auto_question(self, embedding_model):
        """
        질문 리스트를 기반으로 자동으로 질문하고 답변을 받아 엑셀 파일에 저장하는 함수
        :param embedding_model: 임베딩 모델
        """

        df = pd.read_excel("test_automation/qna.xlsx")

        questions_list = df['질의'].tolist()
        correct_answers = df['모범 응답'].tolist()

        for question, correct_answer in zip(questions_list, correct_answers):
            # 거대언어모델이 생성한 응답
            response = self.chain.invoke(question)

            # 코사인 유사도 확인
            temp_model_answer = embedding_model.embed_query(correct_answer)
            temp_response = embedding_model.embed_query(response)
            similarity = self.evaluate.cosine_similarity(temp_model_answer, temp_response)
            print(f"similarity: {similarity}")

            # GPTScore 평가
            gpt_score_response = self.evaluate.gpt_score(question, correct_answer, response)

            # 파일 저장
            self.save_qna_list(question, response, correct_answer, similarity, gpt_score_response)

    def score_calculate(self):

        # research_result 폴더 경로
        folder_path = "research_result"
        output_path = "research_result/output"

        # output 폴더가 존재하지 않으면 생성
        os.makedirs(output_path, exist_ok=True)

        # research_result 폴더 내 모든 파일 처리
        for file_name in os.listdir(folder_path):

            if file_name.startswith('~$'):  # 엑셀 임시 파일 오류 방지
                continue

            file_path = os.path.join(folder_path, file_name)

            print(f"file_path: {file_path}")

            # 엑셀 파일만 처리 (파일 확장자가 .xlsx인 경우)
            if file_name.endswith('.xlsx'):
                print("steste")
                try:
                    # Json 전처리
                    score_preprocessing = GPTScorePreprocessing(
                        file_path,  # 각 파일 경로
                        "Sheet",  # 엑셀 파일의 시트명
                        0,  # model_name 열 순번
                        1  # GPTScore 열 순번
                    )

                    score_preprocessing.run()  # 엑셀 로드 -> Json 추출 -> 정규화

                    # 파일 이름에서 확장자 제거하여 저장 이름 설정
                    save_name = os.path.splitext(file_name)[0]

                    score_preprocessing.save_json(
                        save_path=output_path,
                        save_name=save_name  # 파일 이름 기반으로 저장
                    )

                    print(f"Json 전처리 성공: {file_name}")

                except FileNotFoundError as e:
                    print(f"파일을 찾을 수 없습니다: {file_name} - {e}")
                except ValueError as e:
                    print(f"값 오류 발생: {file_name} - {e}")
                except Exception as e:
                    print(f"오류 발생...{file_name}: {e}")

    def run(self):
        # 자동 질문 및 답변
        self.auto_question(self.embedding_model)

        # GPTScore 점수 계산
        self.score_calculate()


if __name__ == "__main__":
    qna_system = NoRAG()
    embedding_model, chain = qna_system.qna_init()
    experiment = ExperimentAutomation(embedding_model, chain)
    experiment.run()
