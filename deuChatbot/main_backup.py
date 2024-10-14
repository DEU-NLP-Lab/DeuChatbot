import time

from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings import HuggingFaceEmbeddings
from langchain.embeddings import OpenAIEmbeddings
from langchain_community.vectorstores import Chroma
from langchain.chat_models import ChatOpenAI, ChatOllama
from langchain_anthropic import ChatAnthropic
from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler

from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.prompts import ChatPromptTemplate

from langchain.schema.runnable import RunnableLambda, RunnablePassthrough

import os
import shutil
from dotenv import load_dotenv
import openpyxl
from openpyxl import load_workbook

import numpy as np
import pandas as pd

from langchain_community.document_loaders import TextLoader
from langchain.retrievers import BM25Retriever, EnsembleRetriever
from langchain_community.document_transformers import LongContextReorder
from langchain_core.output_parsers import StrOutputParser

from langchain.chains.query_constructor.base import AttributeInfo
from langchain.retrievers.self_query.base import SelfQueryRetriever


def cosine_similarity(a, b):
    """
    코사인 유사도를 확인하기 위한 함수
    """
    dot_product = np.dot(a, b)
    norm_a = np.linalg.norm(a)
    norm_b = np.linalg.norm(b)
    return dot_product / (norm_a * norm_b)


def save_qna_list_v2(q, a, model_answer, model_checker, similarity):
    """
    질의 응답을 엑셀 파일에 추가하는 함수 (중복 질문 제거)
    """

    # 실험 결과 파일 이름
    filename = 'test_automation/qna_list_v2_embedding_openai(text-embedding-ada-002).xlsx'

    # model_checker 값을 모델 이름으로 변환
    model_name = ''
    if model_checker == '1':
        model_name = 'GPT-3.5'
    elif model_checker == '2':
        model_name = 'GPT-4'
    elif model_checker == '3':
        model_name = 'GPT-4o'
    elif model_checker == '4':
        model_name = 'Claude-3-sonnet-20240229'
    elif model_checker == '5':
        model_name = 'Claude-3-opus-20240229'
    elif model_checker == '6':
        model_name = 'Google Gemini-Pro'
    elif model_checker == '7':
        model_name = 'EEVE Korean'
    elif model_checker == '8':
        model_name = 'Llama-3-8B'
    elif model_checker == '9':
        model_name = 'Qwen1.5-14B-Chat'
    elif model_checker == '10':
        model_name = 'Llama-3-MAAL-8B-Instruct-v0.1'

    try:
        # 기존 엑셀 파일 열기
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        # 파일이 없는 경우 새로운 엑셀 파일 생성
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = '거대언어모델'
        sheet['B1'] = '유사도'
        sheet['C1'] = '질의'
        sheet['D1'] = '응답'
        sheet['E1'] = '모범 응답'

    # '모범 응답' 헤더 추가 (파일이 이미 존재하는 경우에도)
    if not isinstance(sheet['E1'].value, str) or sheet['E1'].value != '모범 응답':
        sheet['E1'] = '모범 응답'

    # 기존 질문 목록 가져오기
    existing_questions = set((cell.value, sheet.cell(row=cell.row, column=1).value) for cell in sheet['C'][1:])

    # 중복 질문 확인
    if (q, model_name) not in existing_questions:
        # 새로운 행에 데이터 추가
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=model_name)
        sheet.cell(row=row, column=2, value=similarity)
        sheet.cell(row=row, column=3, value=q)
        sheet.cell(row=row, column=4, value=a)
        sheet.cell(row=row, column=5, value=model_answer)

    # 거대언어모델별로 정렬
    data = list(sheet.values)[1:]
    data.sort(key=lambda x: (x[0], x[1]))

    # 정렬된 데이터로 시트 업데이트
    sheet.delete_rows(2, sheet.max_row)
    for row, row_data in enumerate(data, start=2):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=row, column=col, value=value)

    # 엑셀 파일 저장
    workbook.save(filename)


def save_qna_list(q, a, model_checker, similarity):
    """
    질의 응답을 엑셀 파일에 추가하는 함수 (중복 질문 제거)
    """
    filename = 'test_automation/qna_list_embedding_openai(text-embedding-ada-002).xlsx'

    # model_checker 값을 모델 이름으로 변환
    model_name = ''
    if model_checker == '1':
        model_name = 'GPT-3.5'
    elif model_checker == '2':
        model_name = 'GPT-4'
    elif model_checker == '3':
        model_name = 'GPT-4o'
    elif model_checker == '4':
        model_name = 'Claude-3-sonnet-20240229'
    elif model_checker == '5':
        model_name = 'Claude-3-opus-20240229'
    elif model_checker == '6':
        model_name = 'Google Gemini-Pro'
    elif model_checker == '7':
        model_name = 'EEVE Korean'
    elif model_checker == '8':
        model_name = 'Llama-3-8B'
    elif model_checker == '9':
        model_name = 'Qwen1.5-14B-Chat'
    elif model_checker == '10':
        model_name = 'Llama-3-MAAL-8B-Instruct-v0.1'

    try:
        # 기존 엑셀 파일 열기
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        # 파일이 없는 경우 새로운 엑셀 파일 생성
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = '거대언어모델'
        sheet['B1'] = '유사도'
        sheet['C1'] = '질의'
        sheet['D1'] = '응답'

    # 기존 질문 목록 가져오기
    existing_questions = set((cell.value, sheet.cell(row=cell.row, column=1).value) for cell in sheet['C'][1:])

    # 중복 질문 확인
    if (q, model_name) not in existing_questions:
        # 새로운 행에 데이터 추가
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=model_name)
        sheet.cell(row=row, column=2, value=similarity)
        sheet.cell(row=row, column=3, value=q)
        sheet.cell(row=row, column=4, value=a)

    # 거대언어모델별로 정렬
    data = list(sheet.values)[1:]
    data.sort(key=lambda x: (x[0], x[1]))

    # 정렬된 데이터로 시트 업데이트
    sheet.delete_rows(2, sheet.max_row)
    for row, row_data in enumerate(data, start=2):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=row, column=col, value=value)

    # 엑셀 파일 저장
    workbook.save(filename)


# def chat_llm():
#     """
#     채팅에 사용되는 거대언어모델 생성 함수
#     :return: 답변해주는 거대언어모델
#     """
#     load_dotenv('.env')
#
#     while True:
#         model_check = input(
#             "채팅에 사용할 모델을 고르시오. 고르지 않을 경우 Google Gemini-1.5 Pro 모델을 기본으로 사용합니다.\n"
#             "1: GPT-3.5-turbo\n2: GPT-4-turbo\n"
#             "3: Claude-3-sonnet\n4: Claude-3-opus\n"
#             "5: Google Gemini-Pro\n"
#             "6: EEVE Korean\n7: Llama-3-8B\n8: Qwen1.5-14B-Chat\n9: Llama-3-MAAL-8B-Instruct-v0.1\n\n "
#             "선택 번호 : ")
#
#         if model_check in ['1', '2', '3', '4', '5', '6', '7', '8', '9']:
#             break
#         else:
#             print("잘못된 입력입니다. 1, 2, 3, 4, 5, 6, 7, 8, 9 중 하나를 선택해주세요.\n")
#
#     if model_check == "1":
#         os.getenv("OPENAI_API_KEY")
#         # Retriever 적용
#         llm = ChatOpenAI(
#             model_name="gpt-3.5-turbo",
#             streaming=True, callbacks=[StreamingStdOutCallbackHandler()],
#             temperature=0
#         )
#     elif model_check == "2":
#         os.getenv("OPENAI_API_KEY")
#         # Retriever 적용
#         llm = ChatOpenAI(
#             model_name="gpt-4-turbo",
#             streaming=True, callbacks=[StreamingStdOutCallbackHandler()],
#             temperature=0
#         )
#     elif model_check == "3":
#         os.getenv("ANTHROPIC_API_KEY")
#         # Retriever 적용
#         llm = ChatAnthropic(
#             model_name="claude-3-sonnet-20240229",
#             streaming=True, callbacks=[StreamingStdOutCallbackHandler()],
#             temperature=0
#         )
#     elif model_check == "4":
#         os.getenv("ANTHROPIC_API_KEY")
#         # Retriever 적용
#         llm = ChatAnthropic(
#             model_name="claude-3-opus-20240229",
#             streaming=True, callbacks=[StreamingStdOutCallbackHandler()],
#             temperature=0
#         )
#     elif model_check == "5":
#         os.getenv("GOOGLE_API_KEY")
#
#         llm = ChatGoogleGenerativeAI(
#             model="gemini-1.5-pro-latest",
#             temperature=0
#         )
#     elif model_check == "6":
#         # llm = ChatOllama(model="EEVE-Korean-10.8B:latest")
#         llm = ChatOpenAI(
#             # base_url=os.getenv("LM_URL"),
#             base_url=os.getenv("LM_LOCAL_URL"),
#             api_key="lm-studio",
#             model="teddylee777/EEVE-Korean-Instruct-10.8B-v1.0-gguf",
#             temperature=0,
#             streaming=True,
#             callbacks=[StreamingStdOutCallbackHandler()]
#         )
#         # llm = ChatOllama(model="Llama-3:latest")
#     elif model_check == "7":
#         llm = ChatOpenAI(
#             base_url=os.getenv("LM_URL"),
#             # base_url=os.getenv("LM_LOCAL_URL"),
#             api_key="lm-studio",
#             model="teddylee777/llama-3-8b-it-ko-chang-gguf",
#             temperature=0,
#             streaming=True,
#             callbacks=[StreamingStdOutCallbackHandler()]
#         )
#     elif model_check == "8":
#         llm = ChatOpenAI(
#             base_url=os.getenv("LM_URL"),
#             # base_url=os.getenv("LM_LOCAL_URL"),
#             api_key="lm-studio",
#             model="Qwen/Qwen1.5-14B-Chat-GGUF",
#             temperature=0,
#             streaming=True,
#             callbacks=[StreamingStdOutCallbackHandler()]
#         )
#     elif model_check == "9":
#         llm = ChatOpenAI(
#             base_url=os.getenv("LM_URL"),
#             # base_url=os.getenv("LM_LOCAL_URL"),
#             api_key="lm-studio",
#             model="asiansoul/Llama-3-MAAL-8B-Instruct-v0.1-GGUF",
#             temperature=0,
#             streaming=True,
#             callbacks=[StreamingStdOutCallbackHandler()]
#         )
#
#     if model_check not in ['1', '2', '3', '4', '5', '6', '7', '8', '9']:
#         model_check = '2'  # 디폴트로 구글 제미나이 사용하도록 함
#
#     return llm, model_check

def get_model_info(model_check):
    """
    선택된 모델에 알맞은 정보를 가공하는 함수
    """
    models = {
        "1": {"model_name": "gpt-3.5-turbo", "model_class": ChatOpenAI},
        "2": {"model_name": "gpt-4-turbo", "model_class": ChatOpenAI},
        "3": {"model_name": "gpt-4o", "model_class": ChatOpenAI},
        "4": {"model_name": "claude-3-sonnet-20240229", "model_class": ChatAnthropic},
        "5": {"model_name": "claude-3-opus-20240229", "model_class": ChatAnthropic},
        "6": {"model_name": "gemini-1.5-pro-latest", "model_class": ChatGoogleGenerativeAI},
        "7": {"model_name": "teddylee777/EEVE-Korean-Instruct-10.8B-v1.0-gguf", "model_class": ChatOpenAI,
              "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
        # "base_url": os.getenv("LM_LOCAL_URL"), "api_key": "lm-studio"},
        "8": {"model_name": "teddylee777/llama-3-8b-it-ko-chang-gguf", "model_class": ChatOpenAI,
              "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
        "9": {"model_name": "Qwen/Qwen1.5-14B-Chat-GGUF", "model_class": ChatOpenAI, "base_url": os.getenv("LM_URL"),
              "api_key": "lm-studio"},
        "10": {"model_name": "asiansoul/Llama-3-MAAL-8B-Instruct-v0.1-GGUF", "model_class": ChatOpenAI,
               "base_url": os.getenv("LM_URL"), "api_key": "lm-studio"},
    }

    return models.get(model_check)


def chat_llm():
    """
    채팅에 사용되는 거대언어모델 생성 함수
    :return: 답변해주는 거대언어모델
    """
    load_dotenv('.env')

    while True:
        model_check = input(
            "채팅에 사용할 모델을 고르시오. 고르지 않을 경우 Google Gemini-1.5 Pro 모델을 기본으로 사용합니다.\n"
            "1: GPT-3.5-turbo\n2: GPT-4-turbo\n3: GPT-4o\n"
            "4: Claude-3-sonnet\n5: Claude-3-opus\n"
            "6: Google Gemini-Pro\n"
            "7: EEVE Korean\n8: Llama-3-8B\n9: Qwen1.5-14B-Chat\n10: Llama-3-MAAL-8B-Instruct-v0.1\n\n "
            "선택 번호 : ")

        if model_check in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']:
            break
        else:
            print("잘못된 입력입니다. 1, 2, 3, 4, 5, 6, 7, 8, 9, 10 중 하나를 선택해주세요.\n")

    model_info = get_model_info(model_check)
    model_class = model_info["model_class"]
    model_kwargs = {
        "temperature": 0,
    }

    if model_class == ChatOpenAI:
        model_kwargs.update({
            "model_name": model_info["model_name"],
            "streaming": True,
            "callbacks": [StreamingStdOutCallbackHandler()],
        })
        if "base_url" in model_info:
            model_kwargs["base_url"] = model_info["base_url"]
        if "api_key" in model_info:
            model_kwargs["api_key"] = model_info["api_key"]
    elif model_class == ChatAnthropic:
        model_kwargs.update({
            "model_name": model_info["model_name"],
            "streaming": True,
            "callbacks": [StreamingStdOutCallbackHandler()],
        })
        os.getenv("ANTHROPIC_API_KEY")
    elif model_class == ChatGoogleGenerativeAI:
        os.getenv("GOOGLE_API_KEY")
        model_kwargs.update({
            "model": model_info["model_name"],
        })

    try:
        llm = model_class(**model_kwargs)
    except Exception as e:
        print(f"Error initializing the model: {str(e)}")
        return None, None

    return llm, model_check


def format_docs(docs):
    return "\n\n".join(document.page_content for document in docs)


def reorder_documents(docs):
    # 재정렬
    reordering = LongContextReorder()
    reordered_docs = reordering.transform_documents(docs)
    combined = format_docs(reordered_docs)

    return combined


def db_qna_v3(llm, db, query):
    metadata_field_info = [
        AttributeInfo(
            name="source",
            description="This document contains information about the College Scholastic Ability Test (CSAT) based regular admission process for Dong-Eui University.",
            type="string",
        ),
    ]

    document_content_description = "This document contains information about the College Scholastic Ability Test (CSAT) based regular admission process for Dong-Eui University."

    self_query = SelfQueryRetriever.from_llm(
        llm,
        db,
        document_content_description,
        metadata_field_info,
        verbose=True,
        enable_limit=True,  # 검색 결과 제한 기능을 활성화합니다.
        search_kwargs={'k': 3},
    )

    db = db.as_retriever(
        search_kwargs={'k': 3},
    )

    # 앙상블 retriever를 초기화합니다.
    ensemble_retriever = EnsembleRetriever(
        retrievers=[self_query, db],
        weights=[0.5, 0.5],
        search_type="mmr",
    )

    # ensemble_result = ensemble_retriever.get_relevant_documents(query)
    # self_query_result = self_query.get_relevant_documents(query)
    # db_result = db.get_relevant_documents(query)
    #
    # print("[Ensemble Retriever]\n", ensemble_result, end="\n\n")
    # print("[Self_Query Retriever]\n", self_query_result, end="\n\n")
    # print("[DB Retriever]\n", db_result, end="\n\n")

    prompt = ChatPromptTemplate.from_messages(
        [
            (
                "system",
                """
                You are a specialized AI for question-and-answer tasks.
                You must answer questions based solely on the Context provided.
                For questions about predicting successful applicants, base your answers on data from either the initial successful applicants or the final enrolled students.
                If no Context is provided, you must instruct to inquire at "https://ipsi.deu.ac.kr/main.do".

                Context: {context}
                """,
            ),
            ("human", "Question: {question}"),
        ]
    )

    chain = {
                "context": ensemble_retriever | RunnableLambda(reorder_documents),
                "question": RunnablePassthrough()
            } | prompt | llm | StrOutputParser()

    response = chain.invoke(query)

    if not isinstance(llm, ChatOpenAI):
        print("\n\n{}".format(response))

    return response


def db_qna_v2(llm, bm_db, db, query):
    db = db.as_retriever(
        search_kwargs={'k': 3},
    )
    bm_db.k = 2  # BM25Retriever의 검색 결과 개수를 3로 설정

    # 앙상블 retriever를 초기화합니다.
    ensemble_retriever = EnsembleRetriever(
        retrievers=[bm_db, db],
        weights=[0.4, 0.6],
        search_type="mmr",
    )

    # ensemble_result = ensemble_retriever.get_relevant_documents(query)
    # bm25_result = bm_db.get_relevant_documents(query)
    # db_result = db.get_relevant_documents(query)
    #
    # print("[Ensemble Retriever]\n", ensemble_result, end="\n\n")
    # print("[BM25 Retriever]\n", bm25_result, end="\n\n")
    # print("[db_result Retriever]\n", db_result, end="\n\n")

    prompt = ChatPromptTemplate.from_messages(
        [
            (
                "system",
                """
                You are a specialized AI for question-and-answer tasks.
                You must answer questions based solely on the Context provided.
                For questions about predicting successful applicants, base your answers on data from either the initial successful applicants or the final enrolled students.
                If no Context is provided, you must instruct to inquire at "https://ipsi.deu.ac.kr/main.do".

                Context: {context}
                """,
            ),
            ("human", "Question: {question}"),
        ]
    )

    chain = {
                "context": ensemble_retriever | RunnableLambda(reorder_documents),
                "question": RunnablePassthrough()
            } | prompt | llm | StrOutputParser()

    response = chain.invoke(query)

    if not isinstance(llm, ChatOpenAI):
        print("\n\n{}".format(response))

    return response


def db_qna(llm, db, query, ):
    """
    벡터저장소에서 질문을 검색해서 적절한 답변을 찾아서 답하도록 하는 함수
    :param llm: 거대 언어 모델
    :param db: 벡터스토어
    :param query: 사용자 질문
    :return: 거대언어모델(LLM) 응답 결과
    """

    # docs = db.similarity_search_with_relevance_scores(query, k=3, )
    #
    # for doc in docs:
    #     print("가장 유사한 문서:\n\n {}\n\n".format(doc[0].page_content))
    #     print("문서 유사도:\n {}".format(doc[1]))
    #     print("\n-------------------------")

    db = db.as_retriever(
        # search_kwargs={'k': 3},
        search_type="mmr",
        search_kwargs={'k': 3, 'fetch_k': 10},
        # search_type='similarity_score_threshold',
        # search_kwargs={'k': 3, 'score_threshold': 0.45},
    )

    prompt = ChatPromptTemplate.from_messages(
        [
            (
                "system",
                """
                You are a specialized AI for question-and-answer tasks.
                You must answer questions based solely on the Context provided.
                If no Context is provided, you must instruct to inquire at "https://ipsi.deu.ac.kr/main.do".

                Context: {context}
                """,
            ),
            ("human", "Question: {question}"),
        ]
    )

    chain = {
                "context": db | RunnableLambda(format_docs),
                "question": RunnablePassthrough()
            } | prompt | llm | StrOutputParser()

    response = chain.invoke(query)

    if not isinstance(llm, ChatOpenAI):
        print("\n\n{}".format(response))

    return response


def document_embedding(docs, model, save_directory):
    """
    Embedding 모델을 사용하여 문서 임베딩하여 Chroma 벡터저장소(VectorStore)에 저장하는 함수
    :param model: 임베딩 모델 종류
    :param save_directory: 벡터저장소 저장 경로
    :param docs: 분할된 문서
    :return:
    """

    print("\n잠시만 기다려주세요.\n\n")

    # 벡터저장소가 이미 존재하는지 확인
    if os.path.exists(save_directory):
        shutil.rmtree(save_directory)
        print(f"디렉토리 {save_directory}가 삭제되었습니다.\n")

    if isinstance(model, OpenAIEmbeddings):
        os.getenv("OPENAI_API_KEY")

    print("문서 벡터화를 시작합니다. ")
    db = Chroma.from_documents(docs, model, persist_directory=save_directory)
    print("새로운 Chroma 데이터베이스가 생성되었습니다.\n")

    return db


def document_embedding_v2(docs, model, save_directory):
    """
    Embedding 모델을 사용하여 문서 임베딩하여 Chroma 벡터저장소(VectorStore)에 저장하는 함수 + 앙상블 검색기 기능 추가
    :param model: 임베딩 모델 종류
    :param save_directory: 벡터저장소 저장 경로
    :param docs: 분할된 문서
    :return:
    """

    print("\n잠시만 기다려주세요.\n\n")

    # 벡터저장소가 이미 존재하는지 확인
    if os.path.exists(save_directory):
        shutil.rmtree(save_directory)
        print(f"디렉토리 {save_directory}가 삭제되었습니다.\n")

    if isinstance(model, OpenAIEmbeddings):
        os.getenv("OPENAI_API_KEY")

    print("문서 벡터화를 시작합니다. ")
    db = Chroma.from_documents(docs, model, persist_directory=save_directory)
    bm_db = BM25Retriever.from_documents(
        docs,
    )
    print("새로운 Chroma 데이터베이스가 생성되었습니다.\n")

    return db, bm_db


def c_text_split(corpus):
    """
    CharacterTextSplitter를 사용하여 문서를 분할하도록 하는 함수
    :param corpus: 전처리 완료된 말뭉치
    :return: 분리된 청크
    """

    c_text_splitter = CharacterTextSplitter.from_tiktoken_encoder(
        separator="---",
        chunk_size=1500,
        chunk_overlap=0,
    )

    text_documents = c_text_splitter.split_documents(corpus)

    return text_documents


def docs_load():
    """
    문서 읽는 함수
    """

    loader = TextLoader("corpus/모집요강 전처리 버전 1.txt", encoding="utf-8").load()

    return loader


def auto_question_v2(llm, db, bm_db, model_num, embedding_model):  # llm, db, bm_db, model_num, embedding_model

    df = pd.read_excel("test_automation/qna.xlsx")

    questions_list = df['질의'].tolist()
    model_answers_list = df['모범 응답'].tolist()

    for question, model_answer in zip(questions_list, model_answers_list):
        # response = db_qna(llm, db, question)  # 기본 검색기
        response = db_qna_v2(llm, bm_db, db, question)  # 앙상블 검색기 (키워드 기반 문서 검색, 의미적 유사성 기반 문서 검색)
        # response = db_qna_v3(llm, db, question)  # 앙상블 검색기 (셀프 쿼리 기반 문서 검색, 의미적 유사성 기반 문서 검색,)

        # 코사인 유사도 확인
        temp_model_answer = embedding_model.embed_query(model_answer)
        temp_response = embedding_model.embed_query(response)
        similarity = cosine_similarity(temp_model_answer, temp_response)
        print(f"similarity: {similarity}")

        # 파일 저장
        # save_qna_list(question, response, model_num, similarity)
        save_qna_list_v2(question, response, model_answer, model_num, similarity)
        # time.sleep(20)


def auto_question(llm, db, bm_db, model_num, embedding_model):
    with open("test_automation/question_list.txt", "r", encoding="utf-8") as f:
        question_list = f.read().split("\n")

    for question in question_list:
        # response = db_qna(llm, db, question)  # 기본 검색기
        response = db_qna_v2(llm, bm_db, db, question)  # 앙상블 검색기 (키워드 기반 문서 검색, 의미적 유사성 기반 문서 검색)
        # response = db_qna_v3(llm, db, question)  # 앙상블 검색기 (셀프 쿼리 기반 문서 검색, 의미적 유사성 기반 문서 검색,)

        # 코사인 유사도 확인
        temp_q = embedding_model.embed_query(question)
        temp_a = embedding_model.embed_query(response)
        similarity = cosine_similarity(temp_q, temp_a)

        # 파일 저장
        save_qna_list(question, response, model_num, similarity)


def manual_question(llm, db, bm_db, model_num, embedding_model):
    check = 'Y'  # 0이면 질문 가능
    while check == 'Y' or check == 'y':
        query = input("질문을 입력하세요 : ")
        # response = db_qna(llm, db, query)  # 기본 검색기
        response = db_qna_v2(llm, bm_db, db, query)  # 앙상블 검색기 (키워드 기반 문서 검색, 의미적 유사성 기반 문서 검색)
        # response = db_qna_v3(llm, db, query)  # 앙상블 검색기 (셀프 쿼리 기반 문서 검색, 의미적 유사성 기반 문서 검색,)

        # 코사인 유사도 확인
        temp_q = embedding_model.embed_query(query)
        temp_a = embedding_model.embed_query(response)
        similarity = cosine_similarity(temp_q, temp_a)

        # 파일 저장
        save_qna_list(query, response, model_num, similarity)

        check = input("\n\nY: 계속 질문한다.\nN: 프로그램 종료\n입력: ")


def run():
    """
    챗봇 시작
    Document Load -> Text Splitter -> Ducument Embedding -> VectorStore save -> QA
    """

    # 문서 업로드
    loader = docs_load()

    # 문서 분할
    chunk = c_text_split(loader)

    # 문서 임베딩 및 벡터스토어 저장
    embedding_model_number = input(
        "문서 임베딩에 사용할 임베딩 모델을 고르시오. 고르지 않을 경우 HuggingFaceEmbeddings 모델을 기본으로 사용합니다.\n"
        "1: OpenAIEmbeddings()\n"
        "2: HuggingFaceEmbeddings()\n\n "
        "선택 번호 : ")

    if embedding_model_number == 1:
        # model = OpenAIEmbeddings(model="text-embedding-3-small")
        # model = OpenAIEmbeddings(model="text-embedding-3-large")
        model = OpenAIEmbeddings(model="text-embedding-ada-002")
    else:
        model_name = "jhgan/ko-sroberta-multitask"  # 한국어 모델
        model_kwargs = {'device': 'cpu'}  # cpu를 사용하기 위해 설정
        encode_kwargs = {'normalize_embeddings': True}
        model = HuggingFaceEmbeddings(
            model_name=model_name,
            model_kwargs=model_kwargs,
            encode_kwargs=encode_kwargs
        )

    # db = document_embedding(chunk, model, save_directory="./chroma_db")
    db, bm_db = document_embedding_v2(chunk, model, save_directory="./chroma_db")

    # 채팅에 사용할 거대언어모델(LLM) 선택
    llm, model_num = chat_llm()

    # 정보 검색
    # db = Chroma(persist_directory="./chroma_db", embedding_function=model)

    q_way = input("1. 질의 수동\n2. 질의 자동(실험용)\n\n사용할 방식을 선택하시오(기본값 수동): ")

    if q_way == '2':
        # auto_question(llm, db, bm_db, model_num, model)
        auto_question_v2(llm, db, bm_db, model_num, model)
    else:
        manual_question(llm, db, bm_db, model_num, model)


if __name__ == "__main__":
    os.getenv("LANGCHAIN_TRACING_V2")
    os.getenv("LANGCHAIN_ENDPOINT")
    os.getenv("LANGCHAIN_API_KEY")

    run()
